import openpyxl
import os
from openpyxl import Workbook


class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename

    def check_workbook_existence(self):
        """
        Метод для проверки существования файла рабочей книги.
        Возвращает:
        - True, если файл рабочей книги существует.
        - False, если файл рабочей книги не существует.
        """
        return os.path.exists(self.filename)

    def create_workbook_file(self, name):
        """
        Метод для создания нового файла рабочей книги Excel.

        Аргументы:
        - name: строка, имя нового файла.

        Результат:
        - Создает новый файл рабочей книги Excel с указанным именем.
        - Выводит сообщение о успешном создании файла.

        Примечание:
        Если файл с указанным именем уже существует, он будет перезаписан.
        """
        workbook = Workbook()
        workbook.save(name)
        print(f"Файл {name} успешно создан.")

    def generate_alphabet_sequence(self, N):
        """
        Метод для генерации последовательности из N элементов алфавита (A, B, ..., Z, AA, AB, ..., ZZ, AAA, ...).

        Аргументы:
        - N: целое число, количество элементов в последовательности.

        Результат:
        - Возвращает список, содержащий первые N элементов последовательности алфавита.

        Примечание:
        - Если N <= 26, то список будет содержать однобуквенные значения (A, B, ..., Z).
        - Если N > 26, то будут сгенерированы значения с использованием комбинаций букв (AA, AB, ..., ZZ, AAA, ...).
        """
        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        result = []

        # Добавляем однобуквенные значения
        for char in alphabet:
            result.append(char)

        # Функция для генерации последующих значений
        def generate_next_value(prefix):
            next_char = prefix[-1]
            if next_char != 'Z':
                return prefix[:-1] + alphabet[alphabet.index(next_char) + 1]
            else:
                if len(prefix) == 1:
                    return 'AA'
                else:
                    return generate_next_value(prefix[:-1]) + 'A'

        # Генерируем последующие значения
        while len(result) < N:
            result.append(generate_next_value(result[-1]))

        return result[:N]  # Возвращаем только N элементов

    def remove_sheet_custom(self, sheet_name):
        wb = openpyxl.load_workbook(self.filename)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            wb.save(self.filename)
            print(f"The sheet '{sheet_name}' has been removed.")
        else:
            print(f"The sheet '{sheet_name}' does not exist.")

    def set_column_width(self, filename, sheet_name, column_letters, widths):
        """
        Метод для установки ширины указанных колонок на указанном листе в файле Excel.

        Аргументы:
        - filename: строка, имя файла Excel.
        - sheet_name: строка, имя листа в файле Excel.
        - column_letters: список строк, содержащий буквенные обозначения колонок (например, ['A', 'B', 'C']).
        - widths: список целых чисел, содержащий ширины для соответствующих колонок в column_letters.

        Результат:
        - Нет возвращаемого значения, но ширина указанных колонок на указанном листе файла Excel будет установлена.

        Примечание:
        - Функция изменяет файл Excel на месте, сохраняя результат.
        - Если в column_letters переданы некорректные или несуществующие буквенные обозначения колонок, они будут проигнорированы.
        """

        workbook = openpyxl.load_workbook(filename)
        # Выбираем лист по имени
        sheet = workbook[sheet_name]

        # Устанавливаем ширину указанных колонок
        for column_letter, width in zip(column_letters, widths):
            sheet.column_dimensions[column_letter].width = width

        # Сохраняем изменения в файле Excel
        workbook.save(filename)
        print(f"Ширина колонок {column_letters} на листе {sheet_name} установлена.")

    def get_excel_files(self):
        excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
        return excel_files

    def fill_column(self, list_input, column_number, start_row, end_row):
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb.active

        for i, value in enumerate(list_input, start=start_row):
            if i > end_row:
                break
            if i < start_row + len(list_input):
                sheet.cell(row=i, column=column_number).value = value

        wb.save(self.filename)


# Пример использования:
# excel_handler = ExcelHandler("example.xlsx")
# data = ["Value1", "Value2", "Value3", "Value4", "Value5"]
# excel_handler.fill_column(data, 2, start_row=2, end_row=4)

