import pandas as pd
import openpyxl
from openpyxl import Workbook
import os

class ExcelTable:
    def __init__(self, input_file, output_file):
        self.wb = openpyxl.load_workbook(output_file)
        self.first_sheet = self.wb.sheetnames[0]
        self.input_file = input_file
        self.output_file = output_file

    def read_data_from_F(self):
        # Читаем данные из текстового файла
        with open(self.input_file, 'r') as file:
            data = [line.strip().split() for line in file]

        # Разделяем данные на координаты, среднюю яркость и дисперсию
        coordinates = [(int(row[0]), int(row[1])) for row in data]
        brightness = [float(row[2]) for row in data]
        dispersion = [float(row[3]) for row in data]

        return coordinates, brightness, dispersion

    def workbook_init(self):
        print("Метод инициализации")
        if (self.check_workbook_existence()):
            print(f"Файл {self.output_file} существует.")
        else:
            self.create_exel_file()

    def check_workbook_existence(self):
        return os.path.exists(self.output_file)

    def create_exel_file(self):
        workbook = Workbook()
        workbook.save(self.output_file)
        print(f"Файл {self.output_file} успешно создан.")

    def fill_column(self):
        if self.wb is None:
            print("Workbook is not initialized. Call workbook_init() first.")
            return

        sheet = self.wb.active



        print(sheet["A"][1].value)

        self.wb.save(self.output_file)
        print("Value successfully added to the first cell.")



exel = ExcelTable("outputF3.txt", "example.xlsx")
exel.workbook_init()
exel.fill_column()  # Заполнение столбца

